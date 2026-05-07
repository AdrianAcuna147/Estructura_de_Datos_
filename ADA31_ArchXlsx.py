import os
from openpyxl import load_workbook, Workbook


# -------- MÉTODOS -------- #
def mezcla_directa(lista):
    if len(lista) <= 1:
        return lista

    medio = len(lista) // 2
    izquierda = mezcla_directa(lista[:medio])
    derecha = mezcla_directa(lista[medio:])

    return fusionar(izquierda, derecha)


def fusionar(izquierda, derecha):
    resultado = []
    i = j = 0

    while i < len(izquierda) and j < len(derecha):
        if izquierda[i] < derecha[j]:
            resultado.append(izquierda[i])
            i += 1
        else:
            resultado.append(derecha[j])
            j += 1

    resultado.extend(izquierda[i:])
    resultado.extend(derecha[j:])

    return resultado


# -------- LEER TXT -------- #
def leer_txt(nombre_archivo):
    numeros = []

    with open(nombre_archivo, "r") as archivo:
        contenido = archivo.read()

        for dato in contenido.replace("\n", ",").split(","):
            dato = dato.strip()

            if dato:
                try:
                    numeros.append(int(dato))
                except:
                    pass

    return numeros


# -------- LEER XLSX -------- #
def leer_xlsx(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    numeros = []

    for hoja in wb.sheetnames:
        ws = wb[hoja]

        for fila in ws.iter_rows(values_only=True):
            for celda in fila:
                if celda is not None:
                    try:
                        numeros.append(int(celda))
                    except:
                        pass

    return numeros


# -------- DETECTAR FORMATO -------- #
def leer_archivo(nombre_archivo):
    extension = os.path.splitext(nombre_archivo)[1].lower()

    if extension == ".txt":
        return leer_txt(nombre_archivo)

    elif extension == ".xlsx":
        return leer_xlsx(nombre_archivo)

    return []


# -------- GUARDAR -------- #
def guardar_txt(nombre_archivo, lista):
    with open(nombre_archivo, "w") as archivo:
        archivo.write(",".join(map(str, lista)))


def guardar_xlsx(nombre_archivo, lista):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    for i, numero in enumerate(lista, start=1):
        ws[f"A{i}"] = numero

    wb.save(nombre_archivo)


def guardar_archivo(nombre_archivo, lista):
    extension = os.path.splitext(nombre_archivo)[1].lower()

    if extension == ".txt":
        guardar_txt(nombre_archivo, lista)

    elif extension == ".xlsx":
        guardar_xlsx(nombre_archivo, lista)


# -------- MENÚ -------- #
def menu():
    print("\n====== ORDENAMIENTO DE MÚLTIPLES ARCHIVOS ======")

    archivos = input(
        "Ingresa nombres o rutas separados por comas:\n"
    ).split(",")

    todos_los_numeros = []

    for archivo in archivos:
        archivo = archivo.strip()

        if os.path.exists(archivo):
            numeros = leer_archivo(archivo)

            if numeros:
                print(f"Datos cargados desde {archivo}: {numeros}")
                todos_los_numeros.extend(numeros)

        else:
            print(f"Archivo no encontrado: {archivo}")

    if not todos_los_numeros:
        print("No se encontraron datos válidos.")
        return

    print("\nTodos los datos combinados:")
    print(todos_los_numeros)

    resultado = mezcla_directa(todos_los_numeros)

    print("\nResultado ordenado:")
    print(resultado)

    salida = input(
        "\nNombre del archivo de salida (.txt o .xlsx): "
    )

    guardar_archivo(salida, resultado)

    print(f"Archivo guardado correctamente como {salida}")


# -------- EJECUCIÓN -------- #
menu()
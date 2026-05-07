# ==========================================
# INTERFAZ GRÁFICA PARA ORDENAMIENTO
# - Carga múltiples archivos TXT o XLSX
# - Selecciona método
# - Ordena datos
# - Guarda resultado en archivo nuevo
# ==========================================

import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook


# -------- MÉTODOS DE ORDENAMIENTO -------- #
def intercalacion(lista):
    mitad = len(lista) // 2
    izquierda = sorted(lista[:mitad])
    derecha = sorted(lista[mitad:])

    resultado = []
    i = j = 0

    while i < len(izquierda) and j < len(derecha):
        if izquierda[i] < derecha[j]:
            resultado.append(izquierda[i])
            i += 1
        else:
            resultado.append(derecha[j])
            j += 1

    resultado.extend(izquierda[i:])
    resultado.extend(derecha[j:])

    return resultado


def mezcla_directa(lista):
    if len(lista) <= 1:
        return lista

    medio = len(lista) // 2
    izquierda = mezcla_directa(lista[:medio])
    derecha = mezcla_directa(lista[medio:])

    return fusionar(izquierda, derecha)


def fusionar(izquierda, derecha):
    resultado = []
    i = j = 0

    while i < len(izquierda) and j < len(derecha):
        if izquierda[i] < derecha[j]:
            resultado.append(izquierda[i])
            i += 1
        else:
            resultado.append(derecha[j])
            j += 1

    resultado.extend(izquierda[i:])
    resultado.extend(derecha[j:])

    return resultado


def mezcla_equilibrada(lista):
    if len(lista) <= 1:
        return lista

    mitad = len(lista) // 2
    izquierda = mezcla_equilibrada(lista[:mitad])
    derecha = mezcla_equilibrada(lista[mitad:])

    return combinar(izquierda, derecha)


def combinar(izquierda, derecha):
    resultado = []

    while izquierda and derecha:
        if izquierda[0] < derecha[0]:
            resultado.append(izquierda.pop(0))
        else:
            resultado.append(derecha.pop(0))

    resultado += izquierda
    resultado += derecha

    return resultado


# -------- LECTURA DE ARCHIVOS -------- #
def leer_txt(archivo):
    numeros = []

    with open(archivo, "r") as f:
        contenido = f.read()

        for dato in contenido.replace("\n", ",").split(","):
            dato = dato.strip()

            if dato:
                try:
                    numeros.append(int(dato))
                except:
                    pass

    return numeros


def leer_xlsx(archivo):
    wb = load_workbook(archivo)
    numeros = []

    for hoja in wb.sheetnames:
        ws = wb[hoja]

        for fila in ws.iter_rows(values_only=True):
            for celda in fila:
                if celda is not None:
                    try:
                        numeros.append(int(celda))
                    except:
                        pass

    return numeros


def leer_archivo(archivo):
    extension = os.path.splitext(archivo)[1].lower()

    if extension == ".txt":
        return leer_txt(archivo)

    elif extension == ".xlsx":
        return leer_xlsx(archivo)

    return []


# -------- GUARDAR ARCHIVOS -------- #
def guardar_txt(nombre, lista):
    with open(nombre, "w") as f:
        f.write(",".join(map(str, lista)))


def guardar_xlsx(nombre, lista):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    for i, numero in enumerate(lista, start=1):
        ws[f"A{i}"] = numero

    wb.save(nombre)


def guardar_archivo(nombre, lista):
    extension = os.path.splitext(nombre)[1].lower()

    if extension == ".txt":
        guardar_txt(nombre, lista)

    elif extension == ".xlsx":
        guardar_xlsx(nombre, lista)


# -------- VARIABLES GLOBALES -------- #
archivos_seleccionados = []
datos_totales = []


# -------- FUNCIONES DE INTERFAZ -------- #
def seleccionar_archivos():
    global archivos_seleccionados

    archivos = filedialog.askopenfilenames(
        title="Selecciona archivos",
        filetypes=[("Archivos compatibles", "*.txt *.xlsx")]
    )

    if archivos:
        archivos_seleccionados = archivos
        lista_archivos.delete(0, tk.END)

        for archivo in archivos:
            lista_archivos.insert(tk.END, archivo)


def procesar_archivos():
    global datos_totales
    datos_totales = []

    if not archivos_seleccionados:
        messagebox.showwarning("Advertencia", "Selecciona al menos un archivo.")
        return

    for archivo in archivos_seleccionados:
        datos = leer_archivo(archivo)
        datos_totales.extend(datos)

    if not datos_totales:
        messagebox.showerror("Error", "No se encontraron números válidos.")
        return

    metodo = combo_metodo.get()

    if metodo == "Intercalación":
        resultado = intercalacion(datos_totales)

    elif metodo == "Mezcla Directa":
        resultado = mezcla_directa(datos_totales)

    elif metodo == "Mezcla Equilibrada":
        resultado = mezcla_equilibrada(datos_totales)

    else:
        messagebox.showwarning("Advertencia", "Selecciona un método.")
        return

    resultado_texto.delete("1.0", tk.END)
    resultado_texto.insert(tk.END, str(resultado))

    guardar = messagebox.askyesno("Guardar", "¿Deseas guardar el resultado?")

    if guardar:
        archivo_salida = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[
                ("Archivo TXT", "*.txt"),
                ("Archivo Excel", "*.xlsx")
            ]
        )

        if archivo_salida:
            guardar_archivo(archivo_salida, resultado)
            messagebox.showinfo("Éxito", f"Archivo guardado en:\n{archivo_salida}")


# -------- INTERFAZ -------- #
ventana = tk.Tk()
ventana.title("Ordenamiento de Archivos")
ventana.geometry("800x600")
ventana.config(bg="#f0f0f0")

titulo = tk.Label(
    ventana,
    text="Sistema de Ordenamiento Flexible",
    font=("Arial", 18, "bold"),
    bg="#f0f0f0"
)
titulo.pack(pady=10)

btn_archivos = tk.Button(
    ventana,
    text="Seleccionar Archivos TXT/XLSX",
    command=seleccionar_archivos,
    bg="#4CAF50",
    fg="white",
    font=("Arial", 12)
)
btn_archivos.pack(pady=10)

lista_archivos = tk.Listbox(ventana, width=100, height=8)
lista_archivos.pack(pady=10)

label_metodo = tk.Label(
    ventana,
    text="Selecciona método de ordenamiento:",
    bg="#f0f0f0",
    font=("Arial", 12)
)
label_metodo.pack()

combo_metodo = ttk.Combobox(
    ventana,
    values=[
        "Intercalación",
        "Mezcla Directa",
        "Mezcla Equilibrada"
    ],
    state="readonly",
    width=30
)
combo_metodo.pack(pady=10)

btn_procesar = tk.Button(
    ventana,
    text="Ordenar Datos",
    command=procesar_archivos,
    bg="#2196F3",
    fg="white",
    font=("Arial", 12)
)
btn_procesar.pack(pady=10)

label_resultado = tk.Label(
    ventana,
    text="Resultado:",
    bg="#f0f0f0",
    font=("Arial", 12)
)
label_resultado.pack()

resultado_texto = tk.Text(ventana, height=10, width=90)
resultado_texto.pack(pady=10)

ventana.mainloop()